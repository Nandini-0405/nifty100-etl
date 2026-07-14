import sqlite3
import pandas as pd
import os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
os.makedirs(
    "output",
    exist_ok=True
)

conn = sqlite3.connect(
    "db/nifty100.db"
)

df = pd.read_sql(
    """
    SELECT *
    FROM peer_percentiles
    """,
    conn
)

peer_groups = sorted(
    df[
        "peer_group_name"
    ].dropna().unique()
)

writer = pd.ExcelWriter(
    "output/peer_comparison.xlsx",
    engine="openpyxl"
)

for group in peer_groups:

    group_df = df[
        df["peer_group_name"] == group
    ].copy()

    print(
        f"Exporting {group}..."
    )

    columns = [

        col

        for col in [

            "company_id",

            "year",

            "is_benchmark",

            "return_on_equity_pct",

            "return_on_equity_pct_percentile",

            "net_profit_margin_pct",

            "net_profit_margin_pct_percentile",

            "debt_to_equity",

            "debt_to_equity_percentile",

            "asset_turnover",

            "asset_turnover_percentile",

            "interest_coverage",

            "interest_coverage_percentile",

            "free_cash_flow_cr",

            "free_cash_flow_cr_percentile"

        ]

        if col in group_df.columns

    ]

    export_df = group_df[
        columns
    ]

    export_df.to_excel(

        writer,

        sheet_name=str(group)[:31],

        index=False

    )

writer.close()
wb = load_workbook(
    "output/peer_comparison.xlsx"
)

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

yellow_fill = PatternFill(
    start_color="FFEB9C",
    end_color="FFEB9C",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

gold_fill = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid"
)

for sheet in wb.sheetnames:

    ws = wb[sheet]

    headers = [
        cell.value
        for cell in ws[1]
    ]

    percentile_cols = []

    for i, header in enumerate(
        headers,
        start=1
    ):

        if (
            header
            and "percentile"
            in str(header)
        ):

            percentile_cols.append(i)

    benchmark_col = None

    if "is_benchmark" in headers:

        benchmark_col = (
            headers.index(
                "is_benchmark"
            ) + 1
        )

    for row in range(
        2,
        ws.max_row + 1
    ):

        # Benchmark highlight

        if benchmark_col:

            value = ws.cell(
                row,
                benchmark_col
            ).value

            if value == 1:

                for cell in ws[row]:

                    cell.fill = gold_fill

        # Percentile colors

        for col in percentile_cols:

            value = ws.cell(
                row,
                col
            ).value

            if value is None:

                continue

            if value >= 0.75:

                ws.cell(
                    row,
                    col
                ).fill = green_fill

            elif value <= 0.25:

                ws.cell(
                    row,
                    col
                ).fill = red_fill

            else:

                ws.cell(
                    row,
                    col
                ).fill = yellow_fill

wb.save(
    "output/peer_comparison.xlsx"
)

print(
    "\nFormatting applied!"
)
conn.close()

print(
    "\nCreated:"
)

print(
    "output/peer_comparison.xlsx"
)